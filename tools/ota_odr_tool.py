import streamlit as st
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import sys
import os

sys.path.append(os.path.abspath("core"))
from database import get_connection

TEMPLATE_FILE = "data/templates/template_ota.xlsx"
TABLE_START_ROW = 7
ALL_ODR_FOLDER = Path("ALL ODR")

HEADER_CELLS = {
    "applicant": "B2",
    "city": "F2"
}

COLUMNS = {
    "designation": "B",
    "pn": "C",
    "serial": "D",
    "code_site": "G",
    "status": "I"
}

def load_odr_for_edit(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    st.session_state['applicant'] = ws['B2'].value or ""
    st.session_state['city'] = ws['F2'].value or ""
    
    items = []
    row = TABLE_START_ROW
    while True:
        desig = ws[f"{COLUMNS['designation']}{row}"].value
        if not desig:
            break
        items.append({
            "designation": str(desig),
            "pn": str(ws[f"{COLUMNS['pn']}{row}"].value or ""),
            "serial": str(ws[f"{COLUMNS['serial']}{row}"].value or ""),
            "code_site": str(ws[f"{COLUMNS['code_site']}{row}"].value or ""),
            "status": str(ws[f"{COLUMNS['status']}{row}"].value or "")
        })
        row += 1
    st.session_state.ota_items = items
    wb.close()

def render_tool():
    ALL_ODR_FOLDER.mkdir(exist_ok=True)
    st.title("📋 OTA ODR Filler")
    
    if 'ota_items' not in st.session_state: st.session_state.ota_items = []
    if 'applicant' not in st.session_state: st.session_state['applicant'] = ""
    if 'city' not in st.session_state: st.session_state['city'] = ""
    if 'modify_mode' not in st.session_state: st.session_state['modify_mode'] = False
    if 'modify_filename' not in st.session_state: st.session_state['modify_filename'] = None
    if 'tool_view' not in st.session_state: st.session_state['tool_view'] = "➕ Generate New ODR"

    st.radio("View", ["➕ Generate New ODR", "🔍 Search Existing ODRs"], horizontal=True, label_visibility="collapsed", key="tool_view")

    # ==========================================
    # VIEW 1: GENERATE NEW ODR
    # ==========================================
    if st.session_state.tool_view == "➕ Generate New ODR":
        if st.session_state['modify_mode']:
            st.warning(f"✏️ You are editing: {st.session_state.get('modify_filename')}")

        st.subheader("Header Information")
        col1, col2 = st.columns(2)
        with col1:
            st.text_input("Applicant Name", key="applicant")
        with col2:
            st.selectbox("City", ["Algiers", "Oran", "Other"], key="city")

        st.divider()

        st.subheader("Add Materials")
        st.caption("💡 Tip: The Status (Legacy/Reusable) will auto-fill from the database. Type Serial and press ENTER.")
        
        conn = get_connection()
        df_materials = pd.read_sql_query("SELECT designation, pn, nature FROM ota_materials", conn)
        conn.close()
        designations = df_materials['designation'].tolist()

        with st.form("add_ota_form", clear_on_submit=True):
            col1, col2, col3, col4 = st.columns([3, 2, 2, 2])
            with col1:
                selected_mat = st.selectbox("Désignation", options=designations, key="mat_select")
                
                # --- THE FIX: Auto-grab the Legacy/Reusable status ---
                mat_info = df_materials[df_materials['designation'] == selected_mat].iloc[0]
                auto_status = mat_info['nature'] if mat_info['nature'] else "N/A"
                
            with col2:
                serial = st.text_input("Serial Number", key="serial_input")
            with col3:
                code_site = st.text_input("Code Site", key="site_input")
            with col4:
                # Display the status as read-only so you know what will be saved to Excel
                st.text_input("Status (Auto)", value=auto_status, disabled=True)

            submitted = st.form_submit_button("➕ Add Item to List")
            if submitted:
                final_serial = serial.strip() if serial.strip() else "N/A"
                final_site = code_site.strip() if code_site.strip() else "N/A"
                
                st.session_state.ota_items.append({
                    "designation": selected_mat,
                    "pn": mat_info['pn'],
                    "serial": final_serial,
                    "code_site": final_site,
                    "status": auto_status  # Save the Legacy/Reusable status
                })
                st.rerun()

        if st.session_state.ota_items:
            st.subheader(f"Added Items ({len(st.session_state.ota_items)})")
            df_display = pd.DataFrame(st.session_state.ota_items)
            df_display.index = df_display.index + 1
            st.dataframe(df_display, use_container_width=True)
            
            if st.button("🗑️ Clear All Items"):
                st.session_state.ota_items = []
                st.rerun()
        else:
            st.info("No items added yet.")

        st.divider()

        st.subheader("Generate Output")
        if st.button("🚀 Generate ODR Excel File", type="primary"):
            final_applicant = st.session_state['applicant'].strip() if st.session_state['applicant'].strip() else "N/A"
            final_city = st.session_state['city']
            
            if not st.session_state.ota_items:
                st.error("Please add at least one material.")
            else:
                try:
                    wb = load_workbook(TEMPLATE_FILE)
                    ws = wb.active

                    ws[HEADER_CELLS['applicant']] = final_applicant
                    ws[HEADER_CELLS['city']] = final_city

                    current_row = TABLE_START_ROW
                    for item in st.session_state.ota_items:
                        ws[f"{COLUMNS['designation']}{current_row}"] = item['designation']
                        ws[f"{COLUMNS['pn']}{current_row}"] = item['pn']
                        ws[f"{COLUMNS['serial']}{current_row}"] = item['serial']
                        ws[f"{COLUMNS['code_site']}{current_row}"] = item['code_site']
                        ws[f"{COLUMNS['status']}{current_row}"] = item['status']
                        current_row += 1

                    if st.session_state['modify_mode'] and st.session_state.get('modify_filename'):
                        file_name = st.session_state['modify_filename']
                    else:
                        today_str = date.today().strftime("%Y%m%d")
                        file_name = f"{today_str}_OTA_ODR.xlsx"
                        
                    physical_path = ALL_ODR_FOLDER / file_name
                    wb.save(physical_path)

                    virtual_file = BytesIO()
                    wb.save(virtual_file)
                    wb.close()
                    virtual_file.seek(0)

                    st.success("✅ ODR File generated and saved successfully!")
                    st.download_button(
                        label="⬇️ Download Excel File",
                        data=virtual_file,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # Auto-Reset
                    st.session_state.ota_items = []
                    st.session_state['applicant'] = ""
                    st.session_state['city'] = ""
                    st.session_state['modify_mode'] = False
                    st.session_state['modify_filename'] = None

                except Exception as e:
                    st.error(f"Failed to generate file: {e}")

    # ==========================================
    # VIEW 2: SEARCH EXISTING ODRs
    # ==========================================
    else:
        st.subheader("Search Historical ODR Files")
        st.write("Start typing a Site Code. You can View, Modify, or Delete files.")
        
        valid_exts = ['.xlsx', '.xls']
        all_files = sorted(
            [f.name for f in ALL_ODR_FOLDER.iterdir() if f.suffix.lower() in valid_exts and not f.name.startswith('~$')],
            reverse=True
        )
        
        if not all_files:
            st.info("No ODR files have been generated yet. The 'ALL ODR' folder is empty.")
        else:
            search_query = st.text_input("🔍 Type to search:", placeholder="e.g., 03345 or OTA")
            
            if search_query:
                flexible_query = search_query.lower().lstrip('0')
                if not flexible_query: flexible_query = "0"
                filtered_files = [f for f in all_files if flexible_query in f.lower()]
            else:
                filtered_files = all_files[:10]
                
            display_limit = 15
            files_to_show = filtered_files[:display_limit]
            
            if search_query and not filtered_files:
                st.warning(f"No files found matching '{search_query}'.")
            else:
                for selected_file in files_to_show:
                    file_path = ALL_ODR_FOLDER / selected_file
                    
                    col1, col2, col3, col4 = st.columns([4, 1, 1, 1])
                    with col1:
                        st.write(f"📄 {selected_file}")
                    with col2:
                        is_viewing = st.session_state.get('viewing_file') == selected_file
                        btn_label = "❌ Close" if is_viewing else "👁️ View"
                        if st.button(btn_label, key=f"view_{selected_file}"):
                            st.session_state['viewing_file'] = None if is_viewing else selected_file
                            st.rerun()
                    with col3:
                        if st.button("✏️ Modify", key=f"mod_{selected_file}"):
                            load_odr_for_edit(file_path)
                            st.session_state['modify_mode'] = True
                            st.session_state['modify_filename'] = selected_file
                            st.session_state['tool_view'] = "➕ Generate New ODR"
                            st.rerun()
                    with col4:
                        if st.button("🗑️ Delete", key=f"del_{selected_file}"):
                            st.session_state[f'confirm_del_{selected_file}'] = True
                            st.rerun()
                            
                    if st.session_state.get(f'confirm_del_{selected_file}'):
                        st.warning("Are you sure you want to permanently delete this file?")
                        c1, c2 = st.columns(2)
                        if c1.button("✅ Yes, Delete", key=f"yes_del_{selected_file}"):
                            os.remove(file_path)
                            st.session_state[f'confirm_del_{selected_file}'] = False
                            st.rerun()
                        if c2.button("❌ Cancel", key=f"no_del_{selected_file}"):
                            st.session_state[f'confirm_del_{selected_file}'] = False
                            st.rerun()

                    if st.session_state.get('viewing_file') == selected_file:
                        try:
                            df_preview = pd.read_excel(file_path, skiprows=5, usecols="B,C,D,G,I")
                            df_preview.columns = ["Désignation", "PN", "Serial", "Code Site", "Status"]
                            df_preview = df_preview.dropna(how='all')
                            st.dataframe(df_preview, use_container_width=True, hide_index=True)
                        except Exception as e:
                            st.error(f"Could not read file preview: {e}")
                        st.divider()