import streamlit as st
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import sys
import os

sys.path.append(os.path.abspath("core"))
from database import get_connection, release_connection

TEMPLATE_FILE = "data/templates/template_oa.xlsx"
TABLE_START_ROW = 8
ALL_OA_FOLDER = Path("ALL OA")

def load_oa_for_edit(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    st.session_state['oa_applicant'] = ws['B4'].value or ""
    val_date = ws['E4'].value
    if isinstance(val_date, datetime):
        st.session_state['oa_date'] = val_date.date()
    elif isinstance(val_date, str):
        try: st.session_state['oa_date'] = datetime.strptime(val_date, "%Y-%m-%d").date()
        except: st.session_state['oa_date'] = date.today()
    else:
        st.session_state['oa_date'] = date.today()

    items = []
    row = TABLE_START_ROW
    while True:
        mat_name = ws[f"B{row}"].value
        if not mat_name:
            break
        items.append({
            "code_site": str(ws[f"A{row}"].value or ""),
            "material_name": str(mat_name),
            "code_produit": str(ws[f"C{row}"].value or "N/A"),
            "serial": str(ws[f"D{row}"].value or ""),
            "qty": ws[f"E{row}"].value or 1,
            "etat": str(ws[f"F{row}"].value or "")
        })
        row += 1
    st.session_state.oa_items = items
    wb.close()

def render_tool():
    ALL_OA_FOLDER.mkdir(exist_ok=True)
    st.title("📤 OA Return Filler")
    
    if 'oa_items' not in st.session_state: st.session_state.oa_items = []
    if 'oa_applicant' not in st.session_state: st.session_state['oa_applicant'] = ""
    if 'oa_date' not in st.session_state: st.session_state['oa_date'] = date.today()
    if 'modify_mode' not in st.session_state: st.session_state['modify_mode'] = False
    if 'modify_filename' not in st.session_state: st.session_state['modify_filename'] = None
    if 'tool_view' not in st.session_state: st.session_state['tool_view'] = "➕ Generate New OA"

    st.radio("View", ["➕ Generate New OA", "🔍 Search Existing OAs"], horizontal=True, label_visibility="collapsed", key="tool_view")

    # ==========================================
    # VIEW 1: GENERATE NEW OA
    # ==========================================
    if st.session_state.tool_view == "➕ Generate New OA":
        if st.session_state['modify_mode']:
            st.warning(f"✏️ You are editing: {st.session_state.get('modify_filename')}")

        st.subheader("Header Information")
        col1, col2 = st.columns(2)
        with col1:
            st.text_input("Prepared By (Name)", key="oa_applicant")
        with col2:
            st.date_input("Date", key="oa_date")

        st.divider()

        st.subheader("Add Materials")
        st.caption("💡 Tip: If you type a new material name, it will be saved to the database automatically for next time.")
        
        conn = get_connection()
        df_materials = pd.read_sql_query("SELECT material_name FROM oa_materials", conn)
        release_connection(conn)
        existing_mats = df_materials['material_name'].tolist()
        mat_options = existing_mats + ["➕ Type New Material..."]

        with st.form("add_oa_form", clear_on_submit=True):
            col1, col2, col3, col4, col5 = st.columns([3, 2, 2, 1, 2])
            with col1:
                selected_mat = st.selectbox("Material Name", options=mat_options, key="mat_select")
                new_mat_input = None
                if selected_mat == "➕ Type New Material...":
                    new_mat_input = st.text_input("Enter New Material Name")
            with col2:
                serial = st.text_input("Serial Number", key="serial_input")
            with col3:
                code_site = st.text_input("Code Site", key="site_input")
            with col4:
                qty = st.number_input("Qty", min_value=1, value=1, step=1, key="qty_input")
            with col5:
                etat = st.selectbox("Etat", ["Operationelle", "Faulty"], key="etat_input")

            submitted = st.form_submit_button("➕ Add Item to List")
            if submitted:
                if selected_mat == "➕ Type New Material...":
                    final_mat = new_mat_input.strip() if new_mat_input else ""
                    if final_mat:
                        conn = get_connection()
                        c = conn.cursor()
                        try:
                            c.execute('INSERT INTO oa_materials (material_name) VALUES (%s)', (final_mat,))
                            conn.commit()
                        except sqlite3.IntegrityError:
                            pass 
                        finally:
                            release_connection(conn)
                else:
                    final_mat = selected_mat

                if not final_mat:
                    st.warning("Please enter or select a material name.")
                else:
                    final_serial = serial.strip() if serial.strip() else "N/A"
                    final_site = code_site.strip() if code_site.strip() else "N/A"
                    
                    st.session_state.oa_items.append({
                        "code_site": final_site,
                        "material_name": final_mat,
                        "code_produit": "N/A", 
                        "serial": final_serial,
                        "qty": qty,
                        "etat": etat
                    })
                    st.rerun()

        if st.session_state.oa_items:
            st.subheader(f"Added Items ({len(st.session_state.oa_items)})")
            df_display = pd.DataFrame(st.session_state.oa_items)
            df_display.index = df_display.index + 1
            st.dataframe(df_display, use_container_width=True)
            
            if st.button("🗑️ Clear All Items"):
                st.session_state.oa_items = []
                st.rerun()
        else:
            st.info("No items added yet.")

        st.divider()

        st.subheader("Generate Output")
        if st.button("🚀 Generate OA Excel File", type="primary"):
            final_applicant = st.session_state['oa_applicant'].strip() if st.session_state['oa_applicant'].strip() else "N/A"
            oa_date = st.session_state['oa_date']
            final_site_log = st.session_state.get('site_input', 'N/A')
            
            if not st.session_state.oa_items:
                st.error("Please add at least one material.")
            else:
                try:
                    wb = load_workbook(TEMPLATE_FILE)
                    ws = wb.active

                    ws['B4'] = final_applicant
                    ws['E4'] = str(oa_date)

                    current_row = TABLE_START_ROW
                    for item in st.session_state.oa_items:
                        ws[f"A{current_row}"] = item['code_site']
                        ws[f"B{current_row}"] = item['material_name']
                        ws[f"C{current_row}"] = item['code_produit']
                        ws[f"D{current_row}"] = item['serial']
                        ws[f"E{current_row}"] = item['qty']
                        ws[f"F{current_row}"] = item['etat']
                        current_row += 1

                    if st.session_state['modify_mode'] and st.session_state.get('modify_filename'):
                        file_name = st.session_state['modify_filename']
                    else:
                        today_str = date.today().strftime("%Y%m%d")
                        file_name = f"{today_str}_OA_Return.xlsx"
                        
                    physical_path = ALL_OA_FOLDER / file_name
                    wb.save(physical_path)

                    # --- LOG TO DATABASE ---
                    conn = get_connection()
                    c = conn.cursor()
                    c.execute("INSERT INTO generated_documents (doc_type, client, site_code, file_name) VALUES (%s, %s, %s, %s)",
                              ("OA Return", "Ooredoo", final_site_log, file_name))
                    conn.commit()
                    release_connection(conn)

                    virtual_file = BytesIO()
                    wb.save(virtual_file)
                    wb.close()
                    virtual_file.seek(0)

                    st.success("✅ OA File generated and saved successfully!")
                    st.download_button(
                        label="⬇️ Download Excel File",
                        data=virtual_file,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # Auto-Reset
                    st.session_state.oa_items = []
                    st.session_state['oa_applicant'] = ""
                    st.session_state['oa_date'] = date.today()
                    st.session_state['modify_mode'] = False
                    st.session_state['modify_filename'] = None

                except Exception as e:
                    st.error(f"Failed to generate file: {e}")

    # ==========================================
    # VIEW 2: SEARCH EXISTING OAs
    # ==========================================
    else:
        st.subheader("Search Historical OA Files")
        st.write("Start typing a Site Code or Date. You can View, Modify, or Delete files.")
        
        valid_exts = ['.xlsx', '.xls']
        all_files = sorted(
            [f.name for f in ALL_OA_FOLDER.iterdir() if f.suffix.lower() in valid_exts and not f.name.startswith('~$')],
            reverse=True
        )
        
        if not all_files:
            st.info("No OA files have been generated yet. The 'ALL OA' folder is empty.")
        else:
            search_query = st.text_input("🔍 Type to search:", placeholder="e.g., 20231024 or site code")
            
            if search_query:
                flexible_query = search_query.lower().lstrip('0')
                if not flexible_query: flexible_query = "0"
                filtered_files = [f for f in all_files if flexible_query in f.lower()]
            else:
                filtered_files = all_files[:10]
                
            display_limit = 15
            files_to_show = filtered_files[:display_limit]
            
            if search_query and not filtered_files:
                st.warning(f"No files found matching '{search_query}'.")
            else:
                for selected_file in files_to_show:
                    file_path = ALL_OA_FOLDER / selected_file
                    
                    col1, col2, col3, col4 = st.columns([4, 1, 1, 1])
                    with col1:
                        st.write(f"📄 {selected_file}")
                    with col2:
                        is_viewing = st.session_state.get('viewing_file') == selected_file
                        btn_label = "❌ Close" if is_viewing else "👁️ View"
                        if st.button(btn_label, key=f"view_{selected_file}"):
                            st.session_state['viewing_file'] = None if is_viewing else selected_file
                            st.rerun()
                    with col3:
                        if st.button("✏️ Modify", key=f"mod_{selected_file}"):
                            load_oa_for_edit(file_path)
                            st.session_state['modify_mode'] = True
                            st.session_state['modify_filename'] = selected_file
                            st.session_state['tool_view'] = "➕ Generate New OA"
                            st.rerun()
                    with col4:
                        if st.button("🗑️ Delete", key=f"del_{selected_file}"):
                            st.session_state[f'confirm_del_{selected_file}'] = True
                            st.rerun()
                            
                    if st.session_state.get(f'confirm_del_{selected_file}'):
                        st.warning("Are you sure you want to permanently delete this file?")
                        c1, c2 = st.columns(2)
                        if c1.button("✅ Yes, Delete", key=f"yes_del_{selected_file}"):
                            os.remove(file_path)
                            st.session_state[f'confirm_del_{selected_file}'] = False
                            st.rerun()
                        if c2.button("❌ Cancel", key=f"no_del_{selected_file}"):
                            st.session_state[f'confirm_del_{selected_file}'] = False
                            st.rerun()

                    if st.session_state.get('viewing_file') == selected_file:
                        try:
                            df_preview = pd.read_excel(file_path, skiprows=6, usecols="A,B,C,D,E,F")
                            df_preview.columns = ["Code Site", "Désignation", "Code Produit", "N° Série", "Quantité", "Etat"]
                            df_preview = df_preview.dropna(how='all')
                            st.dataframe(df_preview, use_container_width=True, hide_index=True)
                        except Exception as e:
                            st.error(f"Could not read file preview: {e}")
                        st.divider()