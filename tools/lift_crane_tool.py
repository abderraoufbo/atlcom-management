import streamlit as st
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from io import BytesIO
from pathlib import Path
from datetime import datetime
import sys, os

sys.path.append(os.path.abspath("core"))
from database import get_connection

TEMPLATE_FILE = "data/templates/template_lift_crane.xlsx"
ALL_LC_FOLDER = Path("ALL LIFT_CRANE")

def init_excel_file(client):
    ALL_LC_FOLDER.mkdir(exist_ok=True)
    file_path = ALL_LC_FOLDER / f"Lift_Crane_{client}.xlsx"
    if not file_path.exists():
        # Copy template
        import shutil
        shutil.copy(TEMPLATE_FILE, file_path)
    return file_path

def render_tool():
    client = st.session_state.get('selected_client', 'Common')
    excel_path = init_excel_file(client)
    
    st.title(f"🏗️ LIFT & CRANE Tracker ({client})")
    
    if 'lc_edit_mode' not in st.session_state: st.session_state.lc_edit_mode = False
    if 'lc_edit_site' not in st.session_state: st.session_state.lc_edit_site = None

    # Fetch items from DB
    conn = get_connection()
    df_items = pd.read_sql_query("SELECT item_name, item_code, item_by FROM lift_crane_items", conn)
    conn.close()
    item_names = df_items['item_name'].tolist()

    tab1, tab2 = st.tabs(["📋 View Sites", "➕ Add / Edit Site"])

    # --- TAB 1: VIEW SITES ---
    with tab1:
        wb = load_workbook(excel_path)
        ws = wb["The ATM List"]
        sites_data = []
        for row in range(2, ws.max_row + 1):
            site_id = ws.cell(row, 1).value
            if site_id:
                sites_data.append({
                    "Site ID": site_id, "Activity": ws.cell(row, 2).value,
                    "GF/RT": ws.cell(row, 3).value, "Date": ws.cell(row, 4).value,
                    "SBC": ws.cell(row, 5).value, "Item Name": ws.cell(row, 6).value,
                    "Item Code": ws.cell(row, 7).value, "Item By": ws.cell(row, 8).value,
                    "Qty": ws.cell(row, 9).value, "Work Type": ws.cell(row, 10).value
                })
        wb.close()
        
        if sites_data:
            df_sites = pd.DataFrame(sites_data)
            st.dataframe(df_sites, use_container_width=True, hide_index=True)
            
            st.divider()
            st.subheader("Manage Existing Site")
            site_ids = [s["Site ID"] for s in sites_data]
            selected_site = st.selectbox("Select Site to Edit or Delete", site_ids)
            
            col1, col2 = st.columns(2)
            if col1.button("✏️ Load for Editing"):
                st.session_state.lc_edit_mode = True
                st.session_state.lc_edit_site = selected_site
                st.rerun()
            if col2.button("🗑️ Delete Site"):
                wb = load_workbook(excel_path)
                ws = wb["The ATM List"]
                for r in range(2, ws.max_row + 1):
                    if str(ws.cell(r, 1).value) == str(selected_site):
                        ws.delete_rows(r, 1)
                        break
                if selected_site in wb.sheetnames:
                    wb.remove(wb[selected_site])
                wb.save(excel_path)
                wb.close()
                st.success("Site deleted!")
                st.rerun()
        else:
            st.info("No sites added yet.")

    # --- TAB 2: ADD / EDIT SITE ---
    with tab2:
        edit_mode = st.session_state.lc_edit_mode
        st.subheader("Edit Site" if edit_mode else "Add New Site")
        
        # Default values
        site_id_val, act_val, gf_val, date_val, sbc_val, item_val, qty_val, work_val = "", "", "RT", datetime.now().year, "ALTECOM", "", 1, ""
        
        if edit_mode:
            wb = load_workbook(excel_path)
            ws = wb["The ATM List"]
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, 1).value) == str(st.session_state.lc_edit_site):
                    site_id_val = ws.cell(r, 1).value
                    act_val = ws.cell(r, 2).value or ""
                    gf_val = ws.cell(r, 3).value or "RT"
                    date_val = ws.cell(r, 4).value or datetime.now().year
                    sbc_val = ws.cell(r, 5).value or "ALTECOM"
                    item_val = ws.cell(r, 6).value or ""
                    qty_val = ws.cell(r, 9).value or 1
                    work_val = ws.cell(r, 10).value or ""
                    break
            wb.close()

        with st.form("lc_form"):
            c1, c2, c3 = st.columns(3)
            with c1: site_id = st.text_input("Site ID *", value=site_id_val)
            with c2: activity = st.text_input("Activity", value=act_val)
            with c3: gf_rt = st.text_input("GF/RT", value=gf_val)
            
            c1, c2, c3 = st.columns(3)
            with c1: action_date = st.text_input("Action Date (Year)", value=date_val)
            with c2: sbc = st.text_input("SBC", value=sbc_val)
            with c3: qty = st.number_input("QTY", min_value=1, value=int(qty_val))
            
            item_name = st.selectbox("Item Name *", options=item_names, index=item_names.index(item_val) if item_val in item_names else 0)
            type_work = st.text_input("Type of Work", value=work_val)
            
            uploaded_pics = st.file_uploader("Upload Pictures (Proof)", accept_multiple_files=True, type=['png', 'jpg', 'jpeg'])
            
            if st.form_submit_button("💾 Save Site"):
                if not site_id:
                    st.error("Site ID is required.")
                else:
                    # Get item details
                    item_info = df_items[df_items['item_name'] == item_name].iloc[0]
                    item_code = item_info['item_code']
                    item_by = item_info['item_by']
                    
                    wb = load_workbook(excel_path)
                    ws = wb["The ATM List"]
                    
                    # Find row
                    row_num = None
                    for r in range(2, ws.max_row + 1):
                        if str(ws.cell(r, 1).value) == str(site_id):
                            row_num = r
                            break
                    if not row_num:
                        row_num = ws.max_row + 1

                    # Write data
                    ws.cell(row_num, 1, site_id)
                    ws.cell(row_num, 2, activity)
                    ws.cell(row_num, 3, gf_rt)
                    ws.cell(row_num, 4, action_date)
                    ws.cell(row_num, 5, sbc)
                    ws.cell(row_num, 6, item_name)
                    ws.cell(row_num, 7, item_code)
                    ws.cell(row_num, 8, item_by)
                    ws.cell(row_num, 9, qty)
                    ws.cell(row_num, 10, type_work)
                    
                    # Hyperlink
                    proof_cell = ws.cell(row_num, 12)
                    proof_cell.value = "📷"
                    proof_cell.hyperlink = f"#'{site_id}'!A1"
                    proof_cell.font = Font(color="0000FF", underline="single")
                    
                    # Pictures Sheet
                    if site_id in wb.sheetnames: wb.remove(wb[site_id])
                    ws_pics = wb.create_sheet(str(site_id))
                    back_cell = ws_pics.cell(1, 1)
                    back_cell.value = "🔙 Back to Main List"
                    back_cell.hyperlink = "#'The ATM List'!A1"
                    back_cell.font = Font(color="FFFFFF", bold=True)
                    back_cell.fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
                    
                    if uploaded_pics:
                        start_row = 3
                        for idx, file in enumerate(uploaded_pics):
                            img = PILImage.open(file)
                            max_width = 200
                            ratio = max_width / img.width
                            new_height = int(img.height * ratio)
                            img = img.resize((max_width, new_height), PILImage.Resampling.LANCZOS)
                            
                            img_bytes = BytesIO()
                            img.save(img_bytes, format='PNG')
                            img_bytes.seek(0)
                            
                            xl_img = XLImage(img_bytes)
                            cell_ref = f"B{start_row + idx}"
                            ws_pics.add_image(xl_img, cell_ref)
                            ws_pics.row_dimensions[start_row + idx].height = new_height * 0.75

                    wb.save(excel_path)
                    wb.close()
                    
                    st.session_state.lc_edit_mode = False
                    st.session_state.lc_edit_site = None
                    st.success("Site saved successfully!")
                    st.rerun()