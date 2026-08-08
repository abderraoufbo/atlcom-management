import streamlit as st
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import sys
import os
import xlrd  # For reading old .xls files

# Add core to path to access database
sys.path.append(os.path.abspath("core"))
from database import get_connection, release_connection

# --- Constants ---
TEMPLATE_FILE = "data/templates/template.xlsx"
TABLE_START_ROW = 13
FOOTER_START_ROW = 43
FOOTER_ROW_HEIGHT = 60
ALL_RN_FOLDER = Path("ALL RN")

SITE_CELLS = {
    "date": "D2",
    "magazine": "C3",
    "uop": "C4",
    "code_site": "C5",
    "address": "C6"
}

COLUMNS = {
    "part_number": "A",
    "material_name": "B",
    "nature": "D",
    "serial": "E",
    "quantity": "F",
    "status": "G"
}

# --- Helper to read existing RN for modification ---
def load_rn_for_edit(file_path):
    if file_path.suffix.lower() == '.xls':
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        
        val_date = ws.cell_value(1, 3)
        try:
            if isinstance(val_date, float):
                py_date = xlrd.xldate.xldate_as_datetime(val_date, wb.datemode)
                st.session_state['site_date'] = py_date.date()
            elif isinstance(val_date, str):
                st.session_state['site_date'] = datetime.strptime(val_date, "%Y-%m-%d").date()
            else:
                st.session_state['site_date'] = date.today()
        except:
            st.session_state['site_date'] = date.today()
            
        st.session_state['magazine'] = ws.cell_value(2, 2) or ""
        st.session_state['uop'] = ws.cell_value(3, 2) or ""
        st.session_state['code_site'] = ws.cell_value(4, 2) or ""
        st.session_state['address'] = ws.cell_value(5, 2) or ""
        
        items = []
        row_idx = 12  # Row 13 in Excel is index 12
        while True:
            serial = ws.cell_value(row_idx, 4)  # Column E is index 4
            if not serial:
                break
            items.append({
                "part_number": str(ws.cell_value(row_idx, 0)),
                "material_name": str(ws.cell_value(row_idx, 1)),
                "nature": str(ws.cell_value(row_idx, 3)),
                "serial": str(serial),
                "qty": ws.cell_value(row_idx, 5) if ws.cell_value(row_idx, 5) else 1,
                "status": str(ws.cell_value(row_idx, 6))
            })
            row_idx += 1
            
        st.session_state.mobilis_items = items
        
    else:
        # Modern .xlsx file
        wb = load_workbook(file_path)
        ws = wb.active
        
        val_date = ws['D2'].value
        if isinstance(val_date, datetime):
            st.session_state['site_date'] = val_date.date()
        elif isinstance(val_date, str):
            try:
                st.session_state['site_date'] = datetime.strptime(val_date, "%Y-%m-%d").date()
            except:
                st.session_state['site_date'] = date.today()
        else:
            st.session_state['site_date'] = date.today()
            
        st.session_state['magazine'] = ws['C3'].value or ""
        st.session_state['uop'] = ws['C4'].value or ""
        st.session_state['code_site'] = ws['C5'].value or ""
        st.session_state['address'] = ws['C6'].value or ""
        
        items = []
        row = TABLE_START_ROW
        while True:
            serial = ws[f"{COLUMNS['serial']}{row}"].value
            if not serial:
                break
            items.append({
                "part_number": ws[f"{COLUMNS['part_number']}{row}"].value or "",
                "material_name": ws[f"{COLUMNS['material_name']}{row}"].value or "",
                "nature": ws[f"{COLUMNS['nature']}{row}"].value or "",
                "serial": str(serial),
                "qty": ws[f"{COLUMNS['quantity']}{row}"].value or 1,
                "status": ws[f"{COLUMNS['status']}{row}"].value or ""
            })
            row += 1
            
        st.session_state.mobilis_items = items
        wb.close()

# --- Main Tool Render Function ---
def render_tool():
    ALL_RN_FOLDER.mkdir(exist_ok=True)
    st.title("📄 Mobilis RN Material Form Filler")
    
    # Initialize state variables
    if 'mobilis_items' not in st.session_state: st.session_state.mobilis_items = []
    if 'magazine' not in st.session_state: st.session_state['magazine'] = ""
    if 'uop' not in st.session_state: st.session_state['uop'] = ""
    if 'code_site' not in st.session_state: st.session_state['code_site'] = ""
    if 'address' not in st.session_state: st.session_state['address'] = ""
    if 'site_date' not in st.session_state: st.session_state['site_date'] = date.today()
    if 'modify_mode' not in st.session_state: st.session_state['modify_mode'] = False
    if 'modify_filename' not in st.session_state: st.session_state['modify_filename'] = None
    if 'tool_view' not in st.session_state: st.session_state['tool_view'] = "➕ Generate New RN"

    # Radio button for switching views (allows programmatic switching when Modify is clicked)
    st.radio("View", ["➕ Generate New RN", "🔍 Search Existing RNs"], horizontal=True, label_visibility="collapsed", key="tool_view")

    # ==========================================
    # VIEW 1: GENERATE NEW RN
    # ==========================================
    if st.session_state.tool_view == "➕ Generate New RN":
        if st.session_state['modify_mode']:
            st.warning(f"✏️ You are editing an existing file: {st.session_state.get('modify_filename')}. Click Generate to overwrite it.")

        st.write("Generate your RN Excel files instantly. Data is pulled directly from your central database.")

        st.subheader("Site Information")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.date_input("Date", key="site_date")
            st.text_input("Magazine", key="magazine")
        with col2:
            st.text_input("UOP", key="uop")
            st.text_input("Code Site", key="code_site")
        with col3:
            st.text_input("Address", key="address")

        st.divider()

        st.subheader("Add Materials")
        st.caption("💡 Tip: Type a Serial Number and press ENTER to instantly add the item. Leave blank + Enter to use 'N/A'.")
        
        conn = get_connection()
        df_materials = pd.read_sql_query("SELECT material_name, part_number, nature FROM materials", conn)
        release_connection(conn)
        material_names = df_materials['material_name'].tolist()

        with st.form("add_material_form", clear_on_submit=True):
            col1, col2, col3, col4 = st.columns([3, 2, 1, 2])
            with col1:
                selected_mat = st.selectbox("Material Name", options=material_names, key="mat_select")
            with col2:
                serial = st.text_input("Serial Number", key="serial_input")
            with col3:
                qty = st.number_input("Quantity", min_value=1, value=1, step=1, key="qty_input")
            with col4:
                status = st.selectbox("Status", ["Bon état", "Neuf", "Défectueux"], key="status_input")

            submitted = st.form_submit_button("➕ Add Item to List")
            if submitted:
                final_serial = serial.strip() if serial.strip() else "N/A"
                mat_info = df_materials[df_materials['material_name'] == selected_mat].iloc[0]
                st.session_state.mobilis_items.append({
                    "material_name": selected_mat,
                    "part_number": mat_info['part_number'],
                    "nature": mat_info['nature'],
                    "serial": final_serial,
                    "qty": qty,
                    "status": status
                })
                st.rerun()

        if st.session_state.mobilis_items:
            st.subheader(f"Added Items ({len(st.session_state.mobilis_items)})")
            df_display = pd.DataFrame(st.session_state.mobilis_items)
            df_display.index = df_display.index + 1
            st.dataframe(df_display, use_container_width=True)
            
            if st.button("🗑️ Clear All Items"):
                st.session_state.mobilis_items = []
                st.rerun()
        else:
            st.info("No items added yet.")

        st.divider()

        st.subheader("Generate Output")
        if st.button("🚀 Generate RN Excel File", type="primary"):
            final_magazine = st.session_state['magazine'].strip() if st.session_state['magazine'].strip() else "N/A"
            final_uop = st.session_state['uop'].strip() if st.session_state['uop'].strip() else "N/A"
            final_code_site = st.session_state['code_site'].strip() if st.session_state['code_site'].strip() else "N/A"
            final_address = st.session_state['address'].strip() if st.session_state['address'].strip() else "N/A"
            site_date = st.session_state['site_date']
            
            if not st.session_state.mobilis_items:
                st.error("Please add at least one material.")
            else:
                try:
                    wb = load_workbook(TEMPLATE_FILE)
                    ws = wb.active

                    ws[SITE_CELLS['date']] = str(site_date)
                    ws[SITE_CELLS['magazine']] = final_magazine
                    ws[SITE_CELLS['uop']] = final_uop
                    ws[SITE_CELLS['code_site']] = final_code_site
                    ws[SITE_CELLS['address']] = final_address

                    current_rows = FOOTER_START_ROW - TABLE_START_ROW
                    needed_rows = len(st.session_state.mobilis_items)

                    if needed_rows > current_rows:
                        ws.insert_rows(FOOTER_START_ROW, amount=(needed_rows - current_rows))
                    elif needed_rows < current_rows:
                        ws.delete_rows(TABLE_START_ROW + needed_rows, amount=(current_rows - needed_rows))

                    current_row = TABLE_START_ROW
                    for item in st.session_state.mobilis_items:
                        ws[f"{COLUMNS['part_number']}{current_row}"] = item['part_number']
                        ws[f"{COLUMNS['material_name']}{current_row}"] = item['material_name']
                        ws[f"{COLUMNS['nature']}{current_row}"] = item['nature']
                        ws[f"{COLUMNS['serial']}{current_row}"] = item['serial']
                        ws[f"{COLUMNS['quantity']}{current_row}"] = item['qty']
                        ws[f"{COLUMNS['status']}{current_row}"] = item['status']
                        current_row += 1

                    footer_row = TABLE_START_ROW + needed_rows
                    ws.row_dimensions[footer_row].height = FOOTER_ROW_HEIGHT

                    # Handle file naming & .xls to .xlsx upgrade
                    if st.session_state['modify_mode'] and st.session_state.get('modify_filename'):
                        old_filename = st.session_state['modify_filename']
                        if old_filename.endswith('.xls'):
                            file_name = old_filename.replace('.xls', '.xlsx')
                            os.remove(ALL_RN_FOLDER / old_filename)
                        else:
                            file_name = old_filename
                    else:
                        today_str = date.today().strftime("%Y%m%d")
                        file_name = f"{today_str}_{final_code_site}_RN.xlsx"
                        
                    physical_path = ALL_RN_FOLDER / file_name
                    wb.save(physical_path)

                    # --- LOG TO DATABASE ---
                    conn = get_connection()
                    c = conn.cursor()
                    c.execute("INSERT INTO generated_documents (doc_type, client, site_code, file_name) VALUES (%s, %s, %s, %s)",
                              ("RN", "Mobilis", final_code_site, file_name))
                    conn.commit()
                    release_connection(conn)

                    virtual_file = BytesIO()
                    wb.save(virtual_file)
                    wb.close()
                    virtual_file.seek(0)

                    st.success("✅ File generated and saved successfully!")
                    st.download_button(
                        label="⬇️ Download Excel File",
                        data=virtual_file,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # --- AUTO-RESET THE PAGE ---
                    st.session_state.mobilis_items = []
                    st.session_state['magazine'] = ""
                    st.session_state['uop'] = ""
                    st.session_state['code_site'] = ""
                    st.session_state['address'] = ""
                    st.session_state['site_date'] = date.today()
                    st.session_state['modify_mode'] = False
                    st.session_state['modify_filename'] = None

                except Exception as e:
                    st.error(f"Failed to generate file: {e}")

    # ==========================================
    # VIEW 2: SEARCH EXISTING RNS
    # ==========================================
    else:
        st.subheader("Search Historical RN Files")
        st.write("Start typing a Site Code. You can View, Modify, or Delete files.")
        
        valid_exts = ['.xlsx', '.xls']
        all_files = sorted(
            [f.name for f in ALL_RN_FOLDER.iterdir() if f.suffix.lower() in valid_exts and not f.name.startswith('~$')],
            reverse=True
        )
        
        if not all_files:
            st.info("No RN files have been generated yet. The 'ALL RN' folder is empty.")
        else:
            search_query = st.text_input("🔍 Type to search:", placeholder="e.g., 03345 or 01223")
            
            if search_query:
                flexible_query = search_query.lower().lstrip('0')
                if not flexible_query: flexible_query = "0"
                filtered_files = [f for f in all_files if flexible_query in f.lower()]
            else:
                filtered_files = all_files[:10]
                
            display_limit = 15
            files_to_show = filtered_files[:display_limit]
            
            if search_query and not filtered_files:
                st.warning(f"No files found matching '{search_query}'.")
            else:
                if search_query:
                    st.write(f"Found {len(filtered_files)} file(s). Showing the newest {len(files_to_show)}:")
                else:
                    st.write(f"Showing the {len(files_to_show)} most recent files:")
                    
                for selected_file in files_to_show:
                    file_path = ALL_RN_FOLDER / selected_file
                    
                    if file_path.suffix.lower() == '.xls':
                        mime_type = "application/vnd.ms-excel"
                    else:
                        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    
                    col1, col2, col3, col4 = st.columns([4, 1, 1, 1])
                    with col1:
                        st.write(f"📄 {selected_file}")
                    with col2:
                        is_viewing = st.session_state.get('viewing_file') == selected_file
                        btn_label = "❌ Close" if is_viewing else "👁️ View"
                        if st.button(btn_label, key=f"view_{selected_file}"):
                            st.session_state['viewing_file'] = None if is_viewing else selected_file
                            st.rerun()
                    with col3:
                        if st.button("✏️ Modify", key=f"mod_{selected_file}"):
                            load_rn_for_edit(file_path)
                            st.session_state['modify_mode'] = True
                            st.session_state['modify_filename'] = selected_file
                            st.session_state['tool_view'] = "➕ Generate New RN"
                            st.rerun()
                    with col4:
                        if st.button("🗑️ Delete", key=f"del_{selected_file}"):
                            st.session_state[f'confirm_del_{selected_file}'] = True
                            st.rerun()
                            
                    # Delete Confirmation Logic
                    if st.session_state.get(f'confirm_del_{selected_file}'):
                        st.warning("Are you sure you want to permanently delete this file?")
                        c1, c2 = st.columns(2)
                        if c1.button("✅ Yes, Delete", key=f"yes_del_{selected_file}"):
                            os.remove(file_path)
                            st.session_state[f'confirm_del_{selected_file}'] = False
                            st.rerun()
                        if c2.button("❌ Cancel", key=f"no_del_{selected_file}"):
                            st.session_state[f'confirm_del_{selected_file}'] = False
                            st.rerun()

                    # If viewing, show dataframe
                    if st.session_state.get('viewing_file') == selected_file:
                        try:
                            # Determine engine based on extension
                            engine = 'xlrd' if file_path.suffix.lower() == '.xls' else None
                            df_preview = pd.read_excel(file_path, skiprows=11, usecols="A,B,D,E,F,G", engine=engine)
                            df_preview.columns = ["Part Number", "Material Name", "Nature", "Serial", "Qty", "Status"]
                            df_preview = df_preview.dropna(how='all')
                            st.dataframe(df_preview, use_container_width=True, hide_index=True)
                        except Exception as e:
                            st.error(f"Could not read file preview: {e}")
                        st.divider()