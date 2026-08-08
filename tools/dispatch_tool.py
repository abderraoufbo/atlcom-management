import streamlit as st
import sqlite3
import pandas as pd
import folium
from streamlit_folium import st_folium
from datetime import date, datetime
from openpyxl import load_workbook
from io import BytesIO
from pathlib import Path
import plotly.express as px
import sys, os

sys.path.append(os.path.abspath("core"))
from database import get_connection

DEFAULT_LAT = 36.7538
DEFAULT_LON = 3.0588
TEMPLATE_FILE = "data/templates/template_workplan.xlsx"

ALGERIAN_WILAYAS = {
    "Adrar": (27.8742, -0.2939), "Chlef": (36.1551, 1.3346), "Laghouat": (33.8000, 2.8667),
    "Oum El Bouaghi": (35.8792, 7.1153), "Batna": (35.5500, 6.1741), "Béjaïa": (36.7500, 5.0833),
    "Biskra": (34.8500, 5.7333), "Béchar": (31.6171, -2.2168), "Blida": (36.4700, 2.8300),
    "Bouira": (36.3739, 3.9019), "Tamanrasset": (22.7853, 5.5228), "Tébessa": (35.4044, 8.1208),
    "Tlemcen": (34.8783, -1.3150), "Tiaret": (35.3700, 1.3170), "Tizi Ouzou": (36.7167, 4.0500),
    "Alger": (36.7538, 3.0588), "Djelfa": (34.6700, 3.2500), "Jijel": (36.8000, 5.7667),
    "Sétif": (36.1900, 5.4100), "Saïda": (34.8333, 0.1500), "Skikda": (36.8794, 6.9072),
    "Sidi Bel Abbès": (35.1878, -0.6308), "Annaba": (36.9000, 7.7667), "Guelma": (36.4622, 7.4336),
    "Constantine": (36.3650, 6.6147), "Médéa": (36.2675, 2.7531), "Mostaganem": (35.9311, 0.0892),
    "M'Sila": (35.7053, 4.5417), "Mascara": (35.3966, 0.1403), "Ouargla": (31.9500, 5.3333),
    "Oran": (35.6976, -0.6337), "El Bayadh": (33.6833, 1.0167), "Illizi": (26.4833, 8.4167),
    "Bordj Bou Arréridj": (36.0689, 4.7589), "Boumerdès": (36.7589, 3.4778), "El Tarf": (36.7500, 8.3000),
    "Tindouf": (27.6706, -8.1474), "Tissemsilt": (35.6069, 1.8122), "El Oued": (33.5000, 6.8667),
    "Khenchela": (35.4311, 7.1425), "Souk Ahras": (36.2864, 7.9511), "Tipaza": (36.5897, 2.4483),
    "Mila": (36.4500, 6.2667), "Aïn Defla": (36.2675, 1.9667), "Naâma": (33.2667, -0.3167),
    "Aïn Témouchent": (35.3000, -1.1333), "Ghardaïa": (32.4900, 3.6700), "Relizane": (35.7372, 0.5561),
    "El M'Ghair": (33.6833, 6.1500), "El Meniaa": (30.5800, 2.8800), "Ouled Djellal": (34.4100, 5.0800),
    "Bordj Badji Mokhtar": (27.1500, 5.5000), "Béni Abbès": (31.8500, -2.2000), "Timimoun": (29.2500, 0.2500),
    "Touggourt": (33.1000, 6.0500), "Djanet": (24.5500, 9.4500), "In Salah": (27.2000, 2.4700),
    "In Guezzam": (19.5700, 5.7500)
}

def render_tool():
    st.title("📡 Team Dispatch & Live Tracker")
    
    tab1, tab2, tab3 = st.tabs(["🗺️ Live Map & Dispatcher", "👥 Team Manager", "📈 Analytics & History"])

    # ==========================================
    # TAB 1: MAP & DISPATCHER
    # ==========================================
    with tab1:
        conn = get_connection()
        df_teams = pd.read_sql_query("SELECT * FROM teams", conn)
        conn.close()

        # --- SMART STATUS CALCULATION (APPLIED GLOBALLY) ---
        today = date.today()
        computed_statuses = []
        visual_states = []

        if not df_teams.empty:
            for _, row in df_teams.iterrows():
                start_date_str = row.get('start_date')
                current_state = row.get('state_code') or 'S'
                db_status = row.get('status') or 'Available'
                final_status = db_status
                visual_state = current_state

                if start_date_str and str(start_date_str) not in ['None', 'nan', 'NaT']:
                    try:
                        start_dt = datetime.strptime(str(start_date_str), "%Y-%m-%d").date()
                        if start_dt > today:
                            days_until = (start_dt - today).days
                            final_status = f"Resting (Available in {days_until}d)"
                            visual_state = 'R' # Force Orange
                        else:
                            days_worked = (today - start_dt).days
                            if days_worked >= 30 and current_state != 'R':
                                final_status = f"⚠️ Rest Recommended ({days_worked}d worked)"
                            elif current_state == 'R':
                                final_status = f"Resting (until {row.get('return_to_work_date') or 'N/A'})"
                            else:
                                final_status = f"{db_status} ({days_worked}d worked)"
                    except:
                        pass

                computed_statuses.append(final_status)
                visual_states.append(visual_state)

            df_teams['Computed Status'] = computed_statuses
            df_teams['Visual State'] = visual_states
        
        col1, col2 = st.columns([2, 1])
        
        with col2:
            st.subheader("🔍 Smart Filter")
            req_skill = st.selectbox("Required Skill", ["Any", "MW", "RAN", "Data Center", "Power", "Civil Works"])
            req_status = st.selectbox("Availability", ["Any", "Available", "On-Site", "Resting"])
            
            df_filtered = df_teams.copy()
            if req_skill != "Any":
                df_filtered = df_filtered[df_filtered['skills'].str.contains(req_skill, case=False, na=False)]
            if req_status != "Any":
                df_filtered = df_filtered[df_filtered['Computed Status'].str.contains(req_status, case=False, na=False)]
                
            st.metric("Matching Teams Found", len(df_filtered))
            
            if not df_filtered.empty:
                st.dataframe(df_filtered[['team_name', 'skills', 'Computed Status', 'current_location_name']], use_container_width=True, hide_index=True)
                
                with st.expander("✍️ Assign Mission / Update Status"):
                    team_to_assign = st.selectbox("Select Team", df_filtered['team_name'].tolist())
                    
                    c1, c2 = st.columns(2)
                    with c1: new_wilaya = st.selectbox("Wilaya (For Map Pin)", list(ALGERIAN_WILAYAS.keys()))
                    with c2: new_region = st.text_input("Region (Text for Work Plan)", key="reg_in", placeholder="e.g., Centre Ville")
                    
                    c1, c2 = st.columns(2)
                    with c1: new_project = st.selectbox("Project", ["AT", "ATM", "OTA", "OA", "Other"])
                    with c2: new_state = st.selectbox("State Code", ["W (Working)", "R (Resting)", "S (Stand-by)", "T (On road)", "P (Urgency)"])
                    
                    new_status_notes = st.text_input("Status Notes (Free Text)")
                    new_loc = ""
                    return_date = None
                    
                    if new_state == "W (Working)":
                        new_loc = st.text_input("Site ID / Current Location Name *")
                    if new_state == "R (Resting)":
                        return_date = st.date_input("Planned Return to Work Date", min_value=date.today())
                        
                    if st.button("Update Team Status", type="primary"):
                        if new_state == "W (Working)" and not new_loc:
                            st.error("Please enter the Site ID for the working team.")
                        else:
                            state_letter = new_state[0] 
                            update_lat, update_lon = None, None
                            
                            if new_state == "W (Working)":
                                update_lat, update_lon = ALGERIAN_WILAYAS[new_wilaya]
                                st.info(f"Map pin moved to {new_wilaya}.")
                            elif new_state in ["R (Resting)", "S (Stand-by)"]:
                                team_row = df_teams[df_teams['team_name'] == team_to_assign].iloc[0]
                                update_lat = team_row['home_lat']
                                update_lon = team_row['home_lon']
                                st.info("Map pin moved back to Home Base.")

                            conn = get_connection(); c = conn.cursor()
                            c.execute("""UPDATE teams 
                                         SET wilaya=?, region=?, current_project=?, state_code=?, status_notes=?, current_location_name=?, return_to_work_date=?, current_lat=?, current_lon=? 
                                         WHERE team_name=?""", 
                                      (new_wilaya, new_region, new_project, state_letter, new_status_notes, new_loc, str(return_date) if return_date else None, update_lat, update_lon, team_to_assign))
                            conn.commit(); conn.close()
                            st.success(f"{team_to_assign} updated!")
                            st.rerun()
            else:
                st.info("No teams match your filters.")

        with col1:
            st.subheader("Live Team Map")
            map_center = [DEFAULT_LAT, DEFAULT_LON]
            if not df_teams.empty:
                map_center = [df_teams['current_lat'].fillna(DEFAULT_LAT).mean(), df_teams['current_lon'].fillna(DEFAULT_LON).mean()]
            
            m = folium.Map(location=map_center, zoom_start=6)
            show_homes = st.checkbox("Show Only Home Bases (Hide Current)", value=False)
            
            if not df_teams.empty:
                for _, team in df_teams.iterrows():
                    if not show_homes:
                        lat = team['current_lat'] if pd.notna(team['current_lat']) else DEFAULT_LAT
                        lon = team['current_lon'] if pd.notna(team['current_lon']) else DEFAULT_LON
                        v_state = team['Visual State']
                        
                        color = "blue"
                        if v_state == 'W': color = "green"
                        elif v_state == 'R': color = "orange"
                        elif v_state in ['T', 'P']: color = "red"
                        
                        popup_text = f"<b>{team['team_name']}</b><br>State: {team['Computed Status']}<br>Loc: {team['current_location_name'] or 'N/A'}"
                        folium.Marker([lat, lon], popup=folium.Popup(popup_text, max_width=250), tooltip=f"{team['team_name']} (Current)", icon=folium.Icon(color=color, icon='user', prefix='fa')).add_to(m)
                    else:
                        home_lat = team['home_lat'] if pd.notna(team['home_lat']) else DEFAULT_LAT
                        home_lon = team['home_lon'] if pd.notna(team['home_lon']) else DEFAULT_LON
                        home_popup = f"<b>{team['team_name']} - HOME BASE</b><br>{team['home_location_name'] or 'N/A'}"
                        folium.Marker([home_lat, home_lon], popup=folium.Popup(home_popup, max_width=250), tooltip=f"{team['team_name']} (Home)", icon=folium.Icon(color="cadetblue", icon='home', prefix='fa')).add_to(m)
                
                if show_homes:
                    map_center = [df_teams['home_lat'].fillna(DEFAULT_LAT).mean(), df_teams['home_lon'].fillna(DEFAULT_LON).mean()]
                    m.location = map_center
                
            st_folium(m, width=700, height=500)

    # ==========================================
    # TAB 2: TEAM MANAGER
    # ==========================================
    with tab2:
        st.subheader("Manage Teams")
        
        col_exp1, col_exp2 = st.columns([1, 3])
        with col_exp1:
            st.markdown("##### 📥 Export Work Plan")
            conn = get_connection()
            df_export = pd.read_sql_query("SELECT * FROM teams", conn)
            conn.close()
            
            if not df_export.empty:
                if st.button("Generate Excel File", use_container_width=True):
                    try:
                        if Path(TEMPLATE_FILE).exists():
                            wb = load_workbook(TEMPLATE_FILE)
                        else:
                            from openpyxl import Workbook
                            wb = Workbook()
                        ws = wb.active
                        today_str = date.today().strftime("%Y-%b-%d").upper()
                        ws['A1'] = f"WORK PLAN {today_str}"
                        current_row = 3
                        for _, team in df_export.iterrows():
                            ws[f'B{current_row}'] = team.get('region') or team.get('wilaya') or ""
                            ws[f'C{current_row}'] = team.get('current_project') or ""
                            ws[f'D{current_row}'] = team.get('leader_id') or ""
                            ws[f'E{current_row}'] = team.get('leader_name') or team.get('team_name')
                            ws[f'F{current_row}'] = team.get('status_notes') or ""
                            ws[f'G{current_row}'] = team.get('state_code') or "S"
                            ws[f'H{current_row}'] = team.get('current_location_name') or ""
                            current_row += 1
                        virtual_file = BytesIO()
                        wb.save(virtual_file); wb.close(); virtual_file.seek(0)
                        st.download_button(label="⬇️ Download Excel Report", data=virtual_file, file_name=f"Work_Plan_{today_str}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                        st.success("Excel file mapped successfully!")
                    except Exception as e:
                        st.error(f"Error generating file: {e}")
            else:
                st.warning("No teams to export.")
        
        st.markdown("---")

        with st.expander("➕ Register New Team"):
            with st.form("add_team_form", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1: team_name = st.text_input("Team Name *")
                with c2: leader_id = st.text_input("Team Leader ID *")
                with c3: leader_name = st.text_input("Team Leader Name *")
                c1, c2 = st.columns(2)
                with c1: skills = st.multiselect("Skills", ["MW", "RAN", "Data Center", "Power", "Civil Works"])
                with c2: home_wilaya = st.selectbox("Home Wilaya", list(ALGERIAN_WILAYAS.keys()))
                st.markdown("**Home Base Details**")
                c1, c2 = st.columns(2)
                with c1: home_lat = st.number_input("Home Latitude", value=ALGERIAN_WILAYAS[home_wilaya][0], format="%.4f")
                with c2: home_lon = st.number_input("Home Longitude", value=ALGERIAN_WILAYAS[home_wilaya][1], format="%.4f")
                home_loc = st.text_input("Home Location Name")
                start_date = st.date_input("Start Date of Work", min_value=date(2020,1,1), value=date.today())
                
                if st.form_submit_button("Add Team"):
                    if team_name and leader_id:
                        skills_str = ", ".join(skills)
                        conn = get_connection(); c = conn.cursor()
                        try:
                            c.execute('''INSERT INTO teams 
                                        (team_name, leader_id, leader_name, skills, wilaya, home_lat, home_lon, home_location_name, current_lat, current_lon, current_location_name, status, state_code, start_date)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Available', 'S', ?)''',
                                      (team_name, leader_id, leader_name, skills_str, home_wilaya, home_lat, home_lon, home_loc, home_lat, home_lon, home_loc, str(start_date)))
                            conn.commit(); st.success("Team Added!"); st.rerun()
                        except sqlite3.IntegrityError: st.error("Team already exists.")
                        except Exception as e: st.error(f"Error: {e}")
                        finally: conn.close()

        conn = get_connection()
        df_teams_list = pd.read_sql_query("SELECT team_name FROM teams", conn)
        conn.close()
        
        if not df_teams_list.empty:
            with st.expander("✏️ Edit or Delete Team"):
                edit_team = st.selectbox("Select Team to Edit", df_teams_list['team_name'].tolist())
                conn = get_connection()
                df_edit = pd.read_sql_query("SELECT * FROM teams WHERE team_name=?", conn, params=(edit_team,))
                conn.close()
                
                if not df_edit.empty:
                    team_data = df_edit.iloc[0]
                    with st.form("edit_team_form"):
                        c1, c2, c3 = st.columns(3)
                        with c1: e_leader_id = st.text_input("Leader ID", value=team_data.get('leader_id'))
                        with c2: e_leader_name = st.text_input("Leader Name", value=team_data.get('leader_name'))
                        with c3: 
                            current_skills = [s.strip() for s in str(team_data.get('skills')).split(',') if s.strip()] if team_data.get('skills') else []
                            e_skills = st.multiselect("Skills", ["MW", "RAN", "Data Center", "Power", "Civil Works"], default=current_skills)
                        c1, c2 = st.columns(2)
                        wilaya_list = list(ALGERIAN_WILAYAS.keys())
                        current_wilaya = team_data.get('wilaya')
                        wilaya_idx = wilaya_list.index(current_wilaya) if current_wilaya in wilaya_list else 0
                        with c1: e_home_wilaya = st.selectbox("Home Wilaya", wilaya_list, index=wilaya_idx)
                        with c2: e_home_loc = st.text_input("Home Location Name", value=team_data.get('home_location_name') or "")
                        c1, c2 = st.columns(2)
                        with c1: e_home_lat = st.number_input("Home Latitude", value=float(team_data.get('home_lat')) if pd.notna(team_data.get('home_lat')) else DEFAULT_LAT, format="%.4f")
                        with c2: e_home_lon = st.number_input("Home Longitude", value=float(team_data.get('home_lon')) if pd.notna(team_data.get('home_lon')) else DEFAULT_LON, format="%.4f")
                        e_start_date = st.date_input("Start Date", value=datetime.strptime(team_data.get('start_date'), "%Y-%m-%d").date() if team_data.get('start_date') else date.today())
                        c1, c2 = st.columns(2)
                        with c1: update_btn = st.form_submit_button("🔄 Update Info", use_container_width=True)
                        with c2: pass 
                        
                    if st.button("🗑️ Delete Team Permanently", use_container_width=True):
                        conn = get_connection(); c = conn.cursor()
                        c.execute("DELETE FROM teams WHERE team_name=?", (edit_team,))
                        conn.commit(); conn.close()
                        st.success("Team deleted successfully!"); st.rerun()
                        
                    if update_btn:
                        skills_str = ", ".join(e_skills)
                        conn = get_connection(); c = conn.cursor()
                        c.execute("""UPDATE teams 
                                     SET leader_id=?, leader_name=?, skills=?, wilaya=?, home_location_name=?, home_lat=?, home_lon=?, start_date=? 
                                     WHERE team_name=?""", 
                                  (e_leader_id, e_leader_name, skills_str, e_home_wilaya, e_home_loc, e_home_lat, e_home_lon, str(e_start_date), edit_team))
                        conn.commit(); conn.close()
                        st.success("Team info updated!"); st.rerun()

        st.markdown("---")
        conn = get_connection()
        df = pd.read_sql_query("""SELECT team_name, leader_id, leader_name, skills, wilaya, region, state_code, 
                                  current_location_name, return_to_work_date FROM teams""", conn)
        conn.close()
        st.dataframe(df, use_container_width=True, hide_index=True)

    # ==========================================
    # TAB 3: ANALYTICS & HISTORY
    # ==========================================
    with tab3:
        st.subheader("📈 Team Performance & History")
        
        conn = get_connection()
        df_teams_tab3 = pd.read_sql_query("SELECT team_name FROM teams", conn)
        conn.close()
        
        if df_teams_tab3.empty:
            st.info("Please register teams first to log history.")
        else:
            team_list = df_teams_tab3['team_name'].tolist()
            
            with st.expander("📝 Log Work/Rest History"):
                with st.form("log_history_form", clear_on_submit=True):
                    c1, c2 = st.columns(2)
                    with c1: h_team = st.selectbox("Team", team_list)
                    with c2: h_type = st.selectbox("Activity Type", ["Work", "Rest", "Training", "Travel"])
                    
                    c1, c2, c3 = st.columns(3)
                    with c1: h_loc = st.text_input("Location / Site Name")
                    with c2: h_start = st.date_input("Start Date")
                    with c3: h_end = st.date_input("End Date", min_value=h_start)
                    
                    if st.form_submit_button("Log Activity"):
                        duration = (h_end - h_start).days + 1
                        conn = get_connection(); c = conn.cursor()
                        c.execute("""INSERT INTO team_history 
                                     (team_name, activity_type, location, start_date, end_date, duration_days) 
                                     VALUES (?, ?, ?, ?, ?, ?)""", 
                                  (h_team, h_type, h_loc, str(h_start), str(h_end), duration))
                        conn.commit(); conn.close()
                        st.success(f"Logged {duration} days of {h_type} for {h_team}!"); st.rerun()
            
            st.markdown("---")
            
            conn = get_connection()
            df_history = pd.read_sql_query("SELECT * FROM team_history", conn)
            conn.close()
            
            if df_history.empty:
                st.info("No history logged yet. Use the form above to log past work or rest.")
            else:
                st.write("#### Recent History Logs")
                st.dataframe(df_history[['team_name', 'activity_type', 'location', 'start_date', 'end_date', 'duration_days']].sort_values(by='start_date', ascending=False), use_container_width=True, hide_index=True)
                
                st.markdown("---")
                st.write("#### 📊 Performance Rankings")
                
                df_history['start_date'] = pd.to_datetime(df_history['start_date'])
                df_history['year'] = df_history['start_date'].dt.year
                df_history['month'] = df_history['start_date'].dt.month_name()
                
                c1, c2 = st.columns(2)
                with c1: sel_year = st.selectbox("Select Year", sorted(df_history['year'].unique(), reverse=True))
                with c2: 
                    months = ["All Year"] + sorted(df_history[df_history['year'] == sel_year]['month'].unique(), key=lambda m: datetime.strptime(m, "%B"))
                    sel_month = st.selectbox("Select Month", months)
                
                df_filtered_hist = df_history[df_history['year'] == sel_year]
                if sel_month != "All Year":
                    df_filtered_hist = df_filtered_hist[df_filtered_hist['month'] == sel_month]
                
                if df_filtered_hist.empty:
                    st.warning("No data available for the selected period.")
                else:
                    df_work = df_filtered_hist[df_filtered_hist['activity_type'] == 'Work']
                    if not df_work.empty:
                        df_work_days = df_work.groupby('team_name')['duration_days'].sum().reset_index()
                        df_work_days = df_work_days.sort_values(by='duration_days', ascending=False)
                        st.markdown("##### 🛠️ Total Working Days per Team")
                        fig_work = px.bar(df_work_days, x='team_name', y='duration_days', color='team_name', text='duration_days')
                        fig_work.update_layout(showlegend=False, xaxis_title="Team", yaxis_title="Days Worked")
                        st.plotly_chart(fig_work, use_container_width=True)
                        
                        st.markdown("##### 📉 Total Sites Visited per Team")
                        df_sites = df_work.groupby('team_name')['location'].nunique().reset_index()
                        df_sites = df_sites.sort_values(by='location', ascending=False)
                        fig_sites = px.bar(df_sites, x='team_name', y='location', color='team_name', text='location')
                        fig_sites.update_layout(showlegend=False, xaxis_title="Team", yaxis_title="Number of Sites")
                        st.plotly_chart(fig_sites, use_container_width=True)
                    else:
                        st.info("No 'Work' activities logged for this period.")
                        
                    df_rest = df_filtered_hist[df_filtered_hist['activity_type'] == 'Rest']
                    if not df_rest.empty:
                        df_rest_days = df_rest.groupby('team_name')['duration_days'].sum().reset_index()
                        df_rest_days = df_rest_days.sort_values(by='duration_days', ascending=False)
                        st.markdown("##### 🛌 Total Rest Days per Team")
                        fig_rest = px.bar(df_rest_days, x='team_name', y='duration_days', color='team_name', text='duration_days')
                        fig_rest.update_layout(showlegend=False, xaxis_title="Team", yaxis_title="Days Rested")
                        st.plotly_chart(fig_rest, use_container_width=True)
                    else:
                        st.info("No 'Rest' activities logged for this period.")